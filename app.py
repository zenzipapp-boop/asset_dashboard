"""FastAPI backend for Asset Audit system."""

import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker, Session

# Auto-install dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

from models import Base, Asset, MaintenanceLog

# Database setup
DATABASE_URL = "sqlite:///./assets.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
Base.metadata.create_all(bind=engine)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def ensure_asset_columns():
    required_columns = {
        "disposed": "BOOLEAN DEFAULT 0",
        "disposal_date": "VARCHAR(20)",
        "disposal_reason": "VARCHAR(255)",
        "gst_mode": "VARCHAR(20) DEFAULT 'percent'",
        "gst_place": "VARCHAR(20) DEFAULT 'intra'",
        "cgst_rate": "FLOAT DEFAULT 0",
        "sgst_rate": "FLOAT DEFAULT 0",
        "igst_rate": "FLOAT DEFAULT 0",
        "cgst_amount": "FLOAT DEFAULT 0",
        "sgst_amount": "FLOAT DEFAULT 0",
        "igst_amount": "FLOAT DEFAULT 0",
    }

    with engine.connect() as conn:
        existing = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(assets)").fetchall()}
        for column_name, column_sql in required_columns.items():
            if column_name not in existing:
                conn.exec_driver_sql(f"ALTER TABLE assets ADD COLUMN {column_name} {column_sql}")
        conn.commit()


ensure_asset_columns()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# Pydantic schemas
class AssetCreate(BaseModel):
    asset_name: str
    dop: str
    quantity: int = 1
    si_no: str
    age: str = ""
    disposed: bool = False
    disposal_date: str = ""
    disposal_reason: str = ""
    vendor: str = ""
    value: float = None
    base_price: float = None
    gst_applicable: bool = False
    gst_mode: str = "percent"
    gst_place: str = "intra"
    gst_rate: float = 0
    gst_amount: float = 0
    cgst_rate: float = 0
    sgst_rate: float = 0
    igst_rate: float = 0
    cgst_amount: float = 0
    sgst_amount: float = 0
    igst_amount: float = 0
    location: str = "NA"
    purchased_from: str = "NA"
    department: str = "NA"
    asset_type: str
    status: str = "Functional"


class AssetUpdate(BaseModel):
    asset_name: str = None
    dop: str = None
    quantity: int = None
    si_no: str = None
    age: str = None
    disposed: bool = None
    disposal_date: str = None
    disposal_reason: str = None
    vendor: str = None
    value: float = None
    base_price: float = None
    gst_applicable: bool = None
    gst_mode: str = None
    gst_place: str = None
    gst_rate: float = None
    gst_amount: float = None
    cgst_rate: float = None
    sgst_rate: float = None
    igst_rate: float = None
    cgst_amount: float = None
    sgst_amount: float = None
    igst_amount: float = None
    location: str = None
    purchased_from: str = None
    department: str = None
    asset_type: str = None
    status: str = None


class AssetResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    sl_no: int
    asset_name: str
    dop: str
    quantity: int
    si_no: str
    age: str
    disposed: bool = False
    disposal_date: Optional[str] = None
    disposal_reason: Optional[str] = None
    vendor: str
    value: Optional[float] = None
    base_price: Optional[float] = None
    gst_applicable: bool = False
    gst_mode: str = "percent"
    gst_place: str = "intra"
    gst_rate: float = 0
    gst_amount: float = 0
    cgst_rate: float = 0
    sgst_rate: float = 0
    igst_rate: float = 0
    cgst_amount: float = 0
    sgst_amount: float = 0
    igst_amount: float = 0
    location: str
    purchased_from: str
    department: str
    asset_type: str
    status: str
    created_at: datetime
    updated_at: datetime


class AssetListResponse(BaseModel):
    total: int
    skip: int
    limit: int
    items: list[AssetResponse]


class MaintenanceCreate(BaseModel):
    asset_id: int
    entry_type: str
    date: str
    technician: str
    notes: str = ""
    cost: float = 0
    scrap_value: float = 0
    status: str = "Pending"


class MaintenanceUpdate(BaseModel):
    entry_type: str = None
    date: str = None
    technician: str = None
    notes: str = None
    cost: float = None
    scrap_value: float = None
    status: str = None


class MaintenanceResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    asset_id: int
    entry_type: str
    date: str
    technician: str
    notes: str
    cost: float
    scrap_value: float
    status: str
    created_at: datetime
    updated_at: datetime


# FastAPI app
app = FastAPI(title="Asset Audit System", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Routes ──────────────────────────────────────────────────────────────────────

@app.get("/")
def serve_frontend():
    """Serve the frontend index.html"""
    return FileResponse("index.html")


def build_age(dop: str, end_date: Optional[str] = None) -> str:
    try:
        start = datetime.strptime(dop, "%d/%m/%Y")
        end = datetime.strptime(end_date, "%d/%m/%Y") if end_date else datetime.utcnow()
        months = (end.year - start.year) * 12 + (end.month - start.month)
        if months < 0:
            months = 0
        years, remainder = divmod(months, 12)
        return f"{years} yrs {remainder} mo"
    except Exception:
        return ""


def calc_gst_components(base_price: float, gst_mode: str, gst_place: str, gst_value: float):
    gst_amount = 0.0
    cgst_rate = sgst_rate = igst_rate = 0.0
    cgst_amount = sgst_amount = igst_amount = 0.0

    if gst_place == "vat":
        # VAT: single-rate tax, no CGST/SGST/IGST split
        if gst_mode == "amount":
            gst_amount = float(gst_value or 0)
        else:
            gst_amount = base_price * (float(gst_value or 0) / 100)
    elif gst_mode == "amount":
        gst_amount = float(gst_value or 0)
        if gst_place == "inter":
            igst_amount = gst_amount
        else:
            cgst_amount = gst_amount / 2
            sgst_amount = gst_amount / 2
    else:
        gst_rate = float(gst_value or 0)
        gst_amount = base_price * (gst_rate / 100)
        if gst_place == "inter":
            igst_rate = gst_rate
            igst_amount = gst_amount
        else:
            cgst_rate = gst_rate / 2
            sgst_rate = gst_rate / 2
            cgst_amount = gst_amount / 2
            sgst_amount = gst_amount / 2

    return {
        "gst_amount": gst_amount,
        "cgst_rate": cgst_rate,
        "sgst_rate": sgst_rate,
        "igst_rate": igst_rate,
        "cgst_amount": cgst_amount,
        "sgst_amount": sgst_amount,
        "igst_amount": igst_amount,
    }


SORTABLE_COLS = {
    "sl_no": Asset.sl_no,
    "asset_name": Asset.asset_name,
    "dop": Asset.dop,
    "asset_type": Asset.asset_type,
    "status": Asset.status,
    "value": Asset.value,
    "location": Asset.location,
    "department": Asset.department,
    "vendor": Asset.vendor,
}

@app.get("/api/assets", response_model=AssetListResponse)
def list_assets(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    status: str = Query(None),
    asset_type: str = Query(None),
    location: str = Query(None),
    search: str = Query(None),
    sort_by: str = Query("sl_no"),
    sort_dir: str = Query("asc"),
    db: Session = Depends(get_db),
):
    """List all assets with optional filters, search, and sort."""
    try:
        query = db.query(Asset)

        if status:
            query = query.filter(Asset.status == status)
        if asset_type:
            query = query.filter(Asset.asset_type == asset_type)
        if location:
            query = query.filter(Asset.location == location)
        if search:
            term = f"%{search}%"
            query = query.filter(
                Asset.asset_name.ilike(term) | Asset.si_no.ilike(term)
            )

        sort_col = SORTABLE_COLS.get(sort_by, Asset.sl_no)
        query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col)

        total = query.count()
        assets = query.offset(skip).limit(limit).all()

        return AssetListResponse(
            total=total,
            skip=skip,
            limit=limit,
            items=[AssetResponse.model_validate(a) for a in assets]
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@app.get("/api/assets/{asset_id}")
def get_asset(asset_id: int, db: Session = Depends(get_db)):
    """Get a single asset by ID."""
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return AssetResponse.model_validate(asset)


@app.post("/api/assets", response_model=AssetResponse)
def create_asset(asset: AssetCreate, db: Session = Depends(get_db)):
    """Create a new asset."""
    try:
        # Get next sl_no
        max_sl = db.query(func.max(Asset.sl_no)).scalar() or 0
        new_sl_no = max_sl + 1

        quantity = asset.quantity or 1
        base_price = asset.base_price if asset.gst_applicable else asset.value
        final_value = asset.value
        gst_amount = 0
        cgst_rate = sgst_rate = igst_rate = 0
        cgst_amount = sgst_amount = igst_amount = 0

        if asset.gst_applicable:
            price_basis = float(base_price or 0)
            gst_bits = calc_gst_components(price_basis, asset.gst_mode or "percent", asset.gst_place or "intra", asset.gst_rate if asset.gst_mode == "percent" else asset.gst_amount)
            gst_amount = gst_bits["gst_amount"]
            cgst_rate = gst_bits["cgst_rate"]
            sgst_rate = gst_bits["sgst_rate"]
            igst_rate = gst_bits["igst_rate"]
            cgst_amount = gst_bits["cgst_amount"]
            sgst_amount = gst_bits["sgst_amount"]
            igst_amount = gst_bits["igst_amount"]
            final_value = price_basis + gst_amount

        disposal_date = asset.disposal_date if asset.disposed else None
        age = build_age(asset.dop, disposal_date if asset.disposed else None)
        if asset.age:
            age = asset.age or age

        db_asset = Asset(
            sl_no=new_sl_no,
            asset_name=asset.asset_name,
            dop=asset.dop,
            quantity=quantity,
            si_no=asset.si_no,
            age=age,
            disposed=asset.disposed,
            disposal_date=disposal_date,
            disposal_reason=asset.disposal_reason if asset.disposed else "",
            vendor=asset.vendor,
            value=final_value,
            base_price=base_price,
            gst_applicable=asset.gst_applicable,
            gst_mode=asset.gst_mode,
            gst_place=asset.gst_place,
            gst_rate=asset.gst_rate if asset.gst_applicable else 0,
            gst_amount=gst_amount,
            cgst_rate=cgst_rate,
            sgst_rate=sgst_rate,
            igst_rate=igst_rate,
            cgst_amount=cgst_amount,
            sgst_amount=sgst_amount,
            igst_amount=igst_amount,
            location=asset.location,
            purchased_from=asset.purchased_from,
            department=asset.department,
            asset_type=asset.asset_type,
            status=asset.status,
        )

        db.add(db_asset)
        db.commit()
        db.refresh(db_asset)
        return AssetResponse.model_validate(db_asset)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/assets/{asset_id}", response_model=AssetResponse)
def update_asset(asset_id: int, asset_update: AssetUpdate, db: Session = Depends(get_db)):
    """Update an asset."""
    try:
        db_asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not db_asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        update_data = asset_update.model_dump(exclude_unset=True)
        if "dop" in update_data or "disposed" in update_data or "disposal_date" in update_data:
            dop = update_data.get("dop", db_asset.dop)
            disposed = update_data.get("disposed", db_asset.disposed if hasattr(db_asset, "disposed") else False)
            disposal_date = update_data.get("disposal_date") if disposed else None
            update_data["age"] = build_age(dop, disposal_date if disposed else None)

        if update_data.get("gst_applicable") or update_data.get("gst_mode") or update_data.get("gst_place") or update_data.get("gst_rate") is not None or update_data.get("gst_amount") is not None:
            gst_applicable = update_data.get("gst_applicable", db_asset.gst_applicable)
            gst_mode = update_data.get("gst_mode", db_asset.gst_mode if hasattr(db_asset, "gst_mode") else "percent")
            gst_place = update_data.get("gst_place", db_asset.gst_place if hasattr(db_asset, "gst_place") else "intra")
            base_price = update_data.get("base_price", db_asset.base_price or db_asset.value or 0)
            gst_input = update_data.get("gst_rate", db_asset.gst_rate) if gst_mode == "percent" else update_data.get("gst_amount", db_asset.gst_amount)
            if gst_applicable:
                gst_bits = calc_gst_components(float(base_price or 0), gst_mode, gst_place, float(gst_input or 0))
                update_data.update({
                    "gst_amount": gst_bits["gst_amount"],
                    "cgst_rate": gst_bits["cgst_rate"],
                    "sgst_rate": gst_bits["sgst_rate"],
                    "igst_rate": gst_bits["igst_rate"],
                    "cgst_amount": gst_bits["cgst_amount"],
                    "sgst_amount": gst_bits["sgst_amount"],
                    "igst_amount": gst_bits["igst_amount"],
                    "value": float(base_price or 0) + gst_bits["gst_amount"],
                })

        for field, value in update_data.items():
            setattr(db_asset, field, value)

        db_asset.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(db_asset)
        return AssetResponse.model_validate(db_asset)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/assets/{asset_id}")
def delete_asset(asset_id: int, db: Session = Depends(get_db)):
    """Delete an asset."""
    try:
        db_asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not db_asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        db.delete(db_asset)
        db.commit()
        return {"message": "Asset deleted"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    """Get asset statistics."""
    try:
        total = db.query(func.count(Asset.id)).scalar()
        by_status = {}
        for status in ["Functional", "Non-Functional", "Condemned"]:
            count = db.query(func.count(Asset.id)).filter(Asset.status == status).scalar()
            by_status[status] = count

        by_type = {}
        types = db.query(Asset.asset_type, func.count(Asset.id)).group_by(Asset.asset_type).all()
        for asset_type, count in types:
            by_type[asset_type] = count

        total_value = db.query(func.sum(Asset.value)).scalar() or 0

        return {
            "total_assets": total,
            "by_status": by_status,
            "by_type": by_type,
            "total_value": float(total_value),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@app.get("/api/options")
def get_options(db: Session = Depends(get_db)):
    """Get all unique values for dropdowns."""
    try:
        vendors = [v[0] for v in db.query(Asset.vendor).distinct().all() if v[0]]
        locations = [l[0] for l in db.query(Asset.location).distinct().all() if l[0]]
        departments = [d[0] for d in db.query(Asset.department).distinct().all() if d[0]]
        asset_types = [t[0] for t in db.query(Asset.asset_type).distinct().all() if t[0]]
        asset_names = [n[0] for n in db.query(Asset.asset_name).distinct().all() if n[0]]
        purchased_from = [p[0] for p in db.query(Asset.purchased_from).distinct().all() if p[0]]

        return {
            "vendors": sorted(vendors),
            "locations": sorted(locations),
            "departments": sorted(departments),
            "asset_types": sorted(asset_types),
            "asset_names": sorted(asset_names),
            "purchased_from": sorted(purchased_from),
            "statuses": ["Functional", "Non-Functional", "Condemned"],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@app.post("/api/export-excel")
def export_excel(db: Session = Depends(get_db)):
    """Export all assets to Excel."""
    assets = db.query(Asset).order_by(Asset.sl_no).all()

    if not assets:
        raise HTTPException(status_code=400, detail="No assets to export")

    filepath = "exported_assets.xlsx"

    columns = [
        ("Sl No", "sl_no"),
        ("Name of Asset", "asset_name"),
        ("D.O.P", "dop"),
        ("Quantity", "quantity"),
        ("SI No", "si_no"),
        ("Age of Asset", "age"),
        ("Name of Vendor", "vendor"),
        ("Value", "value"),
        ("Location", "location"),
        ("Purchased From", "purchased_from"),
        ("Department", "department"),
        ("Type of Asset", "asset_type"),
        ("Asset Status", "status"),
    ]

    col_headers = [col[0] for col in columns]
    col_keys = [col[1] for col in columns]
    col_widths = [8, 28, 14, 10, 20, 14, 30, 12, 22, 22, 32, 14, 14]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="DEEAF1")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Register"

    for ci, (hdr, width) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 32

    for ri, asset in enumerate(assets, 2):
        alt = ri % 2 == 0
        for ci, key in enumerate(col_keys, 1):
            val = getattr(asset, key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = border
            if alt:
                cell.fill = alt_fill
            cell.alignment = center if ci in (1, 3, 4, 5, 6, 8, 12, 13) else left

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_headers))}1"
    wb.save(filepath)

    return {"filepath": filepath, "total_rows": len(assets)}


# ── Maintenance Endpoints ──────────────────────────────────────────────

@app.get("/api/maintenance")
def list_maintenance(
    asset_id: int = Query(None),
    status: str = Query(None),
    db: Session = Depends(get_db),
):
    """List maintenance entries with optional filters."""
    query = db.query(MaintenanceLog)

    if asset_id:
        query = query.filter(MaintenanceLog.asset_id == asset_id)
    if status:
        query = query.filter(MaintenanceLog.status == status)

    entries = query.order_by(MaintenanceLog.date.desc()).all()
    return [MaintenanceResponse.model_validate(e) for e in entries]


@app.get("/api/assets/{asset_id}/maintenance")
def get_asset_maintenance(asset_id: int, db: Session = Depends(get_db)):
    """Get maintenance entries for a specific asset."""
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    entries = db.query(MaintenanceLog).filter(MaintenanceLog.asset_id == asset_id).order_by(MaintenanceLog.date.desc()).all()
    return [MaintenanceResponse.model_validate(e) for e in entries]


@app.post("/api/maintenance", response_model=MaintenanceResponse)
def create_maintenance(entry: MaintenanceCreate, db: Session = Depends(get_db)):
    """Create a new maintenance entry."""
    try:
        asset = db.query(Asset).filter(Asset.id == entry.asset_id).first()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        db_entry = MaintenanceLog(
            asset_id=entry.asset_id,
            entry_type=entry.entry_type,
            date=entry.date,
            technician=entry.technician,
            notes=entry.notes,
            cost=entry.cost,
            scrap_value=entry.scrap_value,
            status=entry.status,
        )

        db.add(db_entry)
        db.commit()
        db.refresh(db_entry)
        return MaintenanceResponse.model_validate(db_entry)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/maintenance/{entry_id}", response_model=MaintenanceResponse)
def update_maintenance(entry_id: int, entry_update: MaintenanceUpdate, db: Session = Depends(get_db)):
    """Update a maintenance entry."""
    try:
        db_entry = db.query(MaintenanceLog).filter(MaintenanceLog.id == entry_id).first()
        if not db_entry:
            raise HTTPException(status_code=404, detail="Maintenance entry not found")

        update_data = entry_update.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_entry, field, value)

        db_entry.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(db_entry)
        return MaintenanceResponse.model_validate(db_entry)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/maintenance/{entry_id}")
def delete_maintenance(entry_id: int, db: Session = Depends(get_db)):
    """Delete a maintenance entry."""
    try:
        db_entry = db.query(MaintenanceLog).filter(MaintenanceLog.id == entry_id).first()
        if not db_entry:
            raise HTTPException(status_code=404, detail="Maintenance entry not found")

        db.delete(db_entry)
        db.commit()
        return {"message": "Maintenance entry deleted"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
